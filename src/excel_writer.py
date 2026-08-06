"""
excel_writer.py
----------------
Writes the final report as a styled .xlsx file (replaces the old plain
CSV output). Layout:

    Username | DisplayName | <Account 1> | <Account 2> | ... | <Account N> | DaysInactive

- One column PER DISTINCT ACCOUNT found across all users (dynamic - if
  there are 4 accounts total, you get 4 account columns).
- A user's cell under an account column contains that user's permission
  set(s) for that specific account. Multiple permission sets in one cell
  are separated with a real line break (Alt+Enter equivalent), not commas.
- Users with no access to a given account simply have a blank cell there.
- Header row, alternating row banding, borders, frozen header, and
  auto-filter are applied for readability.

Required package: openpyxl (installed automatically via requirements.txt)
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    FIXED_COLUMNS_LEFT,
    FIXED_COLUMNS_RIGHT_BASE,
    PERMISSION_SET_LINE_BREAK,
    HEADER_FILL_COLOR,
    HEADER_FONT_COLOR,
    HEADER_FONT_SIZE,
    BAND_FILL_COLOR,
    BORDER_COLOR,
    ACCOUNT_COLUMN_MIN_WIDTH,
    ACCOUNT_COLUMN_MAX_WIDTH,
    USERNAME_COLUMN_WIDTH,
    DISPLAYNAME_COLUMN_WIDTH,
    DAYS_INACTIVE_COLUMN_WIDTH,
    DATA_ROW_HEIGHT,
    DATE_FORMAT,
    NOT_CHECKED_LABEL,
    ACTIVE_LABEL,
    INACTIVE_LABEL,
    UNKNOWN_LABEL,
)

ACTIVE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # green
INACTIVE_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # red
UNKNOWN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")   # amber


def _format_account_cell(acct):
    """
    Builds the multi-line cell content for one user's one account column:
        AdministratorAccess
        PowerUserAccess
        Last active: 2026-07-20 (18 days ago)
    """
    lines = sorted(acct["permission_sets"])

    if not acct["checked"]:
        lines.append(f"[{NOT_CHECKED_LABEL}]")
    elif acct["last_active"] is not None:
        date_str = acct["last_active"].strftime(DATE_FORMAT)
        lines.append(f"Last active: {date_str} ({acct['days_inactive']}d ago)")
    else:
        # checked, but no sign-in event found in the lookback window
        lines.append(f"[{acct['days_inactive']}]")

    return PERMISSION_SET_LINE_BREAK.join(lines)


def write_excel(grouped, account_labels, output_path, thresholds_days):
    """
    grouped: OrderedDict from data_merger.merge_user_data(), enriched by
             apply_account_activity() and compute_overall_activity().
    account_labels: sorted list of every distinct account label, from
             data_merger.get_all_account_labels().
    output_path: full path to the .xlsx file to write.
    thresholds_days: list of ints, e.g. [30, 60, 90] - one summary column
             per threshold is added (Active / Inactive / Unknown).
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Inactive Users"

    threshold_headers = [f"Active within {n}d?" for n in thresholds_days]
    headers = FIXED_COLUMNS_LEFT + account_labels + FIXED_COLUMNS_RIGHT_BASE + threshold_headers
    ws.append(headers)

    # ---- Header styling ----
    header_fill = PatternFill(start_color=HEADER_FILL_COLOR, end_color=HEADER_FILL_COLOR, fill_type="solid")
    header_font = Font(color=HEADER_FONT_COLOR, bold=True, size=HEADER_FONT_SIZE)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # ---- Borders ----
    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    band_fill = PatternFill(start_color=BAND_FILL_COLOR, end_color=BAND_FILL_COLOR, fill_type="solid")
    data_align = Alignment(wrap_text=True, vertical="top", horizontal="left")

    # ---- Data rows ----
    row_idx = 2
    n_account_cols = len(account_labels)
    n_threshold_cols = len(thresholds_days)

    for user_id, record in grouped.items():
        row = [record["UserName"], record["DisplayName"]]

        for label in account_labels:
            acct = record["accounts"].get(label)
            row.append(_format_account_cell(acct) if acct else "")

        overall_last_active = record.get("OverallLastActive")
        row.append(overall_last_active.strftime(DATE_FORMAT) if overall_last_active else "")
        overall_days = record.get("OverallDaysInactive")
        row.append(overall_days if overall_days is not None else "No activity found in any checked account")

        threshold_values = [record.get(f"Active_{n}d", UNKNOWN_LABEL) for n in thresholds_days]
        row.extend(threshold_values)

        ws.append(row)

        is_banded = (row_idx % 2 == 0)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = data_align
            if is_banded:
                cell.fill = band_fill

        # Color-code the Active/Inactive/Unknown threshold columns so a
        # user's status is visible at a glance without reading the text.
        first_threshold_col = 1 + 2 + n_account_cols + len(FIXED_COLUMNS_RIGHT_BASE)
        for i, value in enumerate(threshold_values):
            cell = ws.cell(row=row_idx, column=first_threshold_col + i)
            if value == ACTIVE_LABEL:
                cell.fill = ACTIVE_FILL
            elif value == INACTIVE_LABEL:
                cell.fill = INACTIVE_FILL
            elif value == UNKNOWN_LABEL:
                cell.fill = UNKNOWN_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[row_idx].height = DATA_ROW_HEIGHT
        row_idx += 1

    # ---- Column widths ----
    ws.column_dimensions[get_column_letter(1)].width = USERNAME_COLUMN_WIDTH
    ws.column_dimensions[get_column_letter(2)].width = DISPLAYNAME_COLUMN_WIDTH

    for i, label in enumerate(account_labels, start=3):
        col_letter = get_column_letter(i)
        width = max(ACCOUNT_COLUMN_MIN_WIDTH, min(ACCOUNT_COLUMN_MAX_WIDTH, len(label) + 4))
        ws.column_dimensions[col_letter].width = width

    # "Overall Last Active" / "Overall Days Inactive" columns
    overall_start = 3 + n_account_cols
    for offset in range(len(FIXED_COLUMNS_RIGHT_BASE)):
        ws.column_dimensions[get_column_letter(overall_start + offset)].width = DAYS_INACTIVE_COLUMN_WIDTH + 4

    # "Active within Nd?" threshold columns
    threshold_start = overall_start + len(FIXED_COLUMNS_RIGHT_BASE)
    for offset in range(n_threshold_cols):
        ws.column_dimensions[get_column_letter(threshold_start + offset)].width = DAYS_INACTIVE_COLUMN_WIDTH

    # ---- Usability niceties ----
    ws.freeze_panes = "A2"          # keep header visible while scrolling
    ws.auto_filter.ref = ws.dimensions  # dropdown filters on every column
    ws.row_dimensions[1].height = 22

    wb.save(output_path)
    return output_path
